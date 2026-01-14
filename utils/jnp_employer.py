#TEST NAMES
# T1.02 'Find jobs' on JnP Dashboard
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
# from robot.libraries.BuiltIn import BuiltIn
# pip install fuzzywuzzy
# NLP based similarity search library
def match_company_names(company_name1, company_name2, threshold):
    """
    Match two names using fuzzy string matching.
    """
    similarity = fuzz.token_set_ratio(company_name1.lower(), company_name2.lower())
    print('similarity:', similarity)
    # return similarity>=threshold
    return similarity

# c1='Amazon Development Center U.S., Inc.'
# c2='ADC'
# ratio1=fuzz.ratio(c1,c2)
# print(ratio1)
# ratio2=fuzz.partial_ratio(c1,c2)
# print(ratio2)
# ratio3=fuzz.token_set_ratio(c1,c2)
# print(ratio3)
# ratio4=fuzz.token_sort_ratio(c1,c2)
# print(ratio4)

# T2.04 Verification of closing a job and checking with Job-Id
def delete_job_from_db(closed_job_id):
    import mysql.connector
    conn=mysql.connector.connect(
        host="65.109.17.58",
        user="jnp_prod_user",
        password="?Jf03r0d7",
        database="jobsnprofiles_2022"
    )

    cursor=conn.cursor()
    sql="Delete from jnp_jobs where id= %s AND contactemail='automationtest@adamitcorp.com'"
    # sql1 = "DELETE FROM jnp_jobs WHERE contactemail = 'automationtest@adamitcorp.com'"
    job_id=(closed_job_id,)
    try:
        cursor.execute(sql, job_id)
        rowcount = cursor.rowcount
        if rowcount==1:
            # Execute the DELETE query with parameterized value
            # Commit the transaction
            conn.commit()
        # rowcount = cursor.rowcount
    except mysql.connector.Error as err:
        return f"Error in deleting job with id {closed_job_id} from DB: {err}"
        # Rollback in case there is any error
        # conn.rollback()
    finally:
        cursor.close()
        conn.close()
    return f'{cursor.rowcount} Row deleted successfully with job-id {closed_job_id}'

def delete_job_from_solr(closed_job_id):
    import pysolr

    solr_base_url='https://solr.jobsnprofiles.com/solr'
    solr_username='solr'
    solr_password='SolrRocks'
    solr_collection_name_1='jnp_jobs_v2'
    solr_collection_name_2='jnp_jobs_v4_1'
    solr_url_1 = f"{solr_base_url}/{solr_collection_name_1}"
    solr_url_2 = f"{solr_base_url}/{solr_collection_name_2}"
    solr_1 = pysolr.Solr(solr_url_1, always_commit=True, auth=(solr_username, solr_password))
    solr_2 = pysolr.Solr(solr_url_2, always_commit=True, auth=(solr_username, solr_password))

    job_id=closed_job_id

    try:
        solr_1.ping()
        solr_1.delete(id=job_id)
        solr_1.commit()

        solr_2.ping()
        solr_2.delete(id=job_id)
        solr_2.commit()

        return f"'Successfully deleted job with id {closed_job_id} from solr with collection_name '{solr_collection_name_1}' and '{solr_collection_name_2}'"
    except pysolr.SolrError as err:
        return f"Error in deleting job with id {closed_job_id} from solr : {err}"

# closed_job_id=1628310
# result = delete_job_from_db(closed_job_id)
# print(result)



# Test Name- T2.10 Verification of Dashboard, Jobs, Company in a Newly Registered Employer
def delete_company_from_db(company_id,newEmp_company,newEmp_name):
    # Steps to delete a company from the database
    # 1. Collect the uid of the company from 'jnp_companies table and then delete the company. DB: jnp_visa_companies
    # 2. Make the companyid column to '0' in 'jnp_users' table. DB: jobsnprofiles_2022
    import mysql.connector
    conn = mysql.connector.connect(
        host="65.109.17.58",
        user="jnp_prod_user",
        password="?Jf03r0d7",
        database="jnp_visa_companies"
    )
    cursor = conn.cursor()

    # SQL query to fetch the column value for the specific ID
    sql_get_uid = f"SELECT uid FROM jnp_companies WHERE id = %s"
    params = (company_id,)
    try:
        # Execute the SELECT query
        cursor.execute(sql_get_uid, params)

        # Fetch the result
        result = cursor.fetchone()
        if result:
            company_uid=result[0]
            sql_dlt_comp = "Delete from jnp_companies where name= %s AND contactname= %s AND uid= %s"
            params_dlt= (newEmp_company,newEmp_name,company_uid)
            try:
                cursor.execute(sql_dlt_comp, params_dlt)
                rowcount_1 = cursor.rowcount
                if rowcount_1 == 1:
                    conn.commit()
                else:
                    pass
            except mysql.connector.Error as err:
                return f"Error in deleting company with uid {company_uid}: {err}"

            conn_1=mysql.connector.connect(
                host="65.109.17.58",
                user="jnp_prod_user",
                password="?Jf03r0d7",
                database="jobsnprofiles_2022"
            )
            cursor1=conn_1.cursor()
            sql_update_companyid=f'update jnp_users SET companyid=0 where companyid= %s'
            param_1 = (company_id,)
            try:
                cursor1.execute(sql_update_companyid,param_1)
                conn_1.commit()
                rowcount_2=cursor1.rowcount
            except mysql.connector.Error as err:
                return f"Error in updating company with uid {company_uid}: {err}"

            finally:
                cursor1.close()
                conn_1.close()
        else:
            return f"Error in fetching company uid {result}"
    except mysql.connector.Error as err:
        return f"Error in fetching company uid: {err}"

    finally:
        cursor.close()
        conn.close()
    return f"{rowcount_2} Row deleted successfully with company name '{newEmp_company}' with id {company_id} and uid {company_uid}"


# newEmp_comp_url='https://www.lopeds.com'
# company_id= 74858
# newEmp_company='Lapeds'
# newEmp_name='NewEmp Testing'
# result= delete_company_from_db(company_id,newEmp_company,newEmp_name)
# print(result)

def delete_contact_sale_from_db(email,get_name):
    import mysql.connector
    conn=mysql.connector.connect(
        host="65.109.17.58",
        user="jnp_prod_user",
        password="?Jf03r0d7",
        database="jobsnprofiles_2022"
    )
    cur=conn.cursor()
    contact_sale_query="Delete FROM jobsnprofiles_2022.jnp_contactsales where email= %s and firstname='Automation'"
    param1=(email,)
    try:
        cur.execute(contact_sale_query,param1)
        conn.commit()
        rowcount = cur.rowcount
    except mysql.connector.Error as err:
        return f"Error in deleting Contact Sale Request: {err}"

    finally:
        cur.close()
        conn.close()
    return f'{rowcount} Row deleted successfully with email {email} in Contact Sales'

def company_something_went_wrong(newEmp_company,newEmp_name):
    import mysql.connector
    conn = mysql.connector.connect(
        host="65.109.17.58",
        user="jnp_prod_user",
        password="?Jf03r0d7",
        database="jnp_visa_companies"
    )
    cursor=conn.cursor()
    sql_comp_dlt="Delete from jnp_visa_companies.jnp_companies where name= %s AND contactname= %s"
    params=(newEmp_company,newEmp_name,)

    try:
        cursor.execute(sql_comp_dlt, params)
        rowcount = cursor.rowcount
        if rowcount == 1:
            conn.commit()
            # rowcount = cursor.rowcount
        else:
            return f'Company not deleted. Rowcount is {rowcount}'
    except mysql.connector.Error as err:
        return f"Error in deleting company with company name {newEmp_company}: {err}"

    finally:
        cursor.close()
        conn.close()
    return f'{rowcount} Row deleted successfully with company {newEmp_company} from DB'



# result=company_something_went_wrong('Lapeds','NewEmp Testing')
# print(result)
def check_if_company_exists(company_name):
    import mysql.connector
    conn = mysql.connector.connect(
        host="65.109.17.58",
        user="jnp_prod_user",
        password="?Jf03r0d7",
        database="jnp_visa_companies"
    )
    cursor = conn.cursor()
    sql_check_company="SELECT * FROM jnp_visa_companies.jnp_companies where name= %s"
    params=(company_name,)
    try:
        cursor.execute(sql_check_company, params)
        result = cursor.fetchall()  # Fetch all results
        rowcount = len(result)
    except mysql.connector.Error as err:
        return f"Error in fetching company with company name {company_name}: {err}"

    finally:
        cursor.close()
        conn.close()
    # return result[0][0]
    return rowcount

# company_name='Adam It Corp Inc'
# result=check_if_company_exists(company_name)
# print(result)

def delete_emp_from_db(newEmp_email):
    import mysql.connector
    conn = mysql.connector.connect(
        host="65.109.17.58",
        user="jnp_prod_user",
        password="?Jf03r0d7",
        database="jobsnprofiles_2022"
    )
    cursor = conn.cursor()
    sql_check_emp = "Delete FROM jobsnprofiles_2022.jnp_users where username = %s"
    params = (newEmp_email,)
    try:
        cursor.execute(sql_check_emp, params)
        if cursor.rowcount==1:
            conn.commit()
        else:
            pass
        # rowcount = cursor.rowcount
    except mysql.connector.Error as err:
        return f"Error in deleting New EMP with  {newEmp_email}: {err}"

    finally:
        cursor.close()
        conn.close()
    # return result[0][0]
    return f'{cursor.rowcount} Row deleted successfully with new emp mail id:  {newEmp_email}'

# newEmp_email='kehona@lapeds.com'
# print(delete_emp_from_db(newEmp_email))
def delete_comp_from_db(newEmp_company,newEmp_email):
    # Steps to delete a company from the database
    # 1. Collect the uid of the company from 'jnp_companies table and then delete the company. DB: jnp_visa_companies
    # 2. Make the companyid column to '0' in 'jnp_users' table. DB: jobsnprofiles_2022
    import mysql.connector
    conn = mysql.connector.connect(
        host="65.109.17.58",
        user="jnp_prod_user",
        password="?Jf03r0d7",
        database="jnp_visa_companies"
    )
    cursor = conn.cursor()

    # SQL query to fetch the column value for the specific ID
    sql_get_uid = f"SELECT uid FROM jnp_visa_companies.jnp_companies WHERE contactemail = %s"
    params = (newEmp_email,)
    try:
        # Execute the SELECT query
        cursor.execute(sql_get_uid, params)

        # Fetch the result
        result = cursor.fetchone()
        if result:
            company_uid = result[0]
            sql_dlt_comp = "Delete from jnp_companies where name= %s AND contactemail= %s AND uid= %s"
            params_dlt = (newEmp_company, newEmp_email, company_uid)
            try:
                cursor.execute(sql_dlt_comp, params_dlt)
                rowcount_1 = cursor.rowcount
                if rowcount_1 == 1:
                    conn.commit()
                else:
                    pass
            except mysql.connector.Error as err:
                return f"Error in deleting company with uid {company_uid}: {err}"

            conn_1 = mysql.connector.connect(
                host="65.109.17.58",
                user="jnp_prod_user",
                password="?Jf03r0d7",
                database="jobsnprofiles_2022"
            )
            cursor1 = conn_1.cursor()
            sql_update_companyid = f'update jnp_users SET companyid=0 where username= %s'
            param_1 = (newEmp_email,)
            try:
                cursor1.execute(sql_update_companyid, param_1)
                rowcount_2 = cursor1.rowcount
                if rowcount_2==1:
                    conn_1.commit()

            except mysql.connector.Error as err:
                return f"Error in updating company with uid {company_uid}: {err}"

            finally:
                cursor1.close()
                conn_1.close()
        else:
            return f"Error in fetching company uid {result}"
    except mysql.connector.Error as err:
        return f"Error in fetching company uid: {err}"

    finally:
        cursor.close()
        conn.close()
    return f"{rowcount_2} Row deleted successfully with company name '{newEmp_company}' and uid {company_uid}"

# output=delete_comp_from_db('wirelay','ritwik@wirelay.com',)
# print(output)

def read_excel_file(excel_path, sheet_name):
    """
    Read Excel file and return all values from column B as a list.
    Replaces RPA.Excel.Files functionality.
    """
    try:
        import openpyxl
    except ImportError:
        # Fallback to alternative if openpyxl not available
        try:
            import pandas as pd
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
            # Return column B (index 1) values as list, skipping empty cells
            values = [str(val) for val in df.iloc[:, 1].dropna().tolist() if str(val).strip()]
            return values
        except ImportError:
            raise ImportError("Either openpyxl or pandas must be installed. Install with: pip install openpyxl")
    
    # Using openpyxl
    workbook = openpyxl.load_workbook(excel_path)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active  # Use active sheet if specified sheet not found
    
    values = []
    # Read column B (index 2, since openpyxl is 1-indexed)
    for row in sheet.iter_rows(min_row=1, min_col=2, max_col=2, values_only=True):
        if row[0] and str(row[0]).strip():  # Skip empty cells
            values.append(str(row[0]))
    
    workbook.close()
    return values
